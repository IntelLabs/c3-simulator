#!/usr/bin/env python3
# Copyright 2024 Intel Corporation
# SPDX-License-Identifier: MIT

#
# Copyright (C) 2019 Intel Corporation. All rights reserved.
#
# File: process_output.py
#
# Description: Processes output generated by running all the testcases for a CWE.
#
# Author: Salmin Sultana
#
from openpyxl import load_workbook
import pandas as pd

def add_dataframe_to_excelsheet(data_frame, dest_excel, dest_excelsheet, row_val=0, ival=True, hval=True):
    workbook = load_workbook(dest_excel)
    writer = pd.ExcelWriter(dest_excel,engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.workbook = workbook

    data_frame.to_excel(writer, sheet_name=dest_excelsheet, startrow=row_val, index=ival, header=hval)
    writer.close()

    return 0

def write_csv_to_excelsheet(source_csv, dest_excel, dest_excelsheet):
    source_df_1 = pd.read_csv(source_csv,nrows=1)
    source_df_2 = pd.read_csv(source_csv,skiprows=2)

    writer = pd.ExcelWriter(dest_excel)

    source_df_1.to_excel(writer, sheet_name=dest_excelsheet,index=False)
    source_df_2.to_excel(writer, sheet_name=dest_excelsheet,index=False,startrow=2)
    writer.close()
    return 0

def add_csv_to_excelsheet(source_csv, dest_excel, dest_excelsheet):
    source_df_1 = pd.read_csv(source_csv,nrows=1)
    source_df_2 = pd.read_csv(source_csv,skiprows=2)

    workbook = load_workbook(dest_excel)
    writer = pd.ExcelWriter(dest_excel,engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.workbook = workbook

    source_df_1.to_excel(writer, sheet_name=dest_excelsheet,index=False)
    source_df_2.to_excel(writer, sheet_name=dest_excelsheet,index=False,startrow=2)
    writer.close()

    return 0

def summarize_results(output_file):
    workbook = load_workbook(output_file)
    excel_sheet_names=workbook.sheetnames

    num_testcases=[]
    detected_testcases=[]

    for s in excel_sheet_names:
        dataframe=pd.read_excel(output_file,sheet_name=s,header=None,nrows=2)
        num_testcases.append(dataframe[1].values[0])
        detected_testcases.append(dataframe[1].values[1])

    total_testcases = sum(num_testcases)
    total_detected_testcases= sum(detected_testcases)
    detection_rate=round(100 * float(total_detected_testcases) / float(total_testcases), 4)
    print("C3 detection rate: %f" %(detection_rate))

    rdf=pd.DataFrame(data=zip(['C3 detection rate'], [str(detection_rate)+"%"]), columns=None, index=None)
    add_dataframe_to_excelsheet(rdf, output_file, "Summary", 0, False, False)

    headers=['Num of testcases','C3 detected testcases']
    indices=excel_sheet_names.copy()
    indices.append('Total')
    num_testcases.append(total_testcases)
    detected_testcases.append(total_detected_testcases)
    ddf=pd.DataFrame(data=zip(num_testcases, detected_testcases), columns=headers, index=indices)
    add_dataframe_to_excelsheet(ddf, output_file, "Summary",2)

    return 0